"""Finalize and verify generated blind-review workbooks."""

from __future__ import annotations

import argparse
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook


FORBIDDEN_MARKERS = (
    "CAND_",
    "rare_expert_census",
    "weighted_formula_score",
    "honeypot_risk",
    "consistency_score",
    "current_submission_rank",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", default="outputs/annotation_audit")
    return parser.parse_args()


def workbook_xml(path: Path) -> str:
    with ZipFile(path) as archive:
        return b"".join(
            archive.read(name) for name in archive.namelist() if name.endswith(".xml")
        ).decode("utf-8", errors="ignore")


def finalize(path: Path) -> list[str]:
    workbook = load_workbook(path)
    expected_sheets = ["Instructions", "Rubric", "Annotations"]
    if workbook.sheetnames != expected_sheets:
        raise ValueError(f"{path.name}: expected sheets {expected_sheets}, got {workbook.sheetnames}")

    workbook["Instructions"].freeze_panes = "A3"
    workbook["Rubric"].freeze_panes = "A3"
    annotations = workbook["Annotations"]
    annotations.freeze_panes = "B2"
    annotations.auto_filter.ref = "A1:Q201"

    if annotations.max_row != 201 or annotations.max_column != 17:
        raise ValueError(
            f"{path.name}: expected 201x17 annotation region, got "
            f"{annotations.max_row}x{annotations.max_column}"
        )

    headers = [cell.value for cell in annotations[1]]
    expected_tail = ["Relevance 0-5", "Confidence 1-3", "Risk Flag", "Evidence-based Rationale"]
    if headers[-4:] != expected_tail:
        raise ValueError(f"{path.name}: annotation input columns are incorrect")

    profile_keys = [annotations.cell(row=row, column=1).value for row in range(2, 202)]
    if len(set(profile_keys)) != 200 or any(not str(key).startswith("PROFILE-") for key in profile_keys):
        raise ValueError(f"{path.name}: profile keys are missing or duplicated")

    validations = list(annotations.data_validations.dataValidation)
    if len(validations) < 3:
        raise ValueError(f"{path.name}: expected at least three annotation validations")

    workbook.save(path)
    return profile_keys


def main() -> None:
    args = parse_args()
    output_dir = Path(args.output_dir)
    paths = [output_dir / f"annotation_reviewer_{reviewer}.xlsx" for reviewer in (1, 2)]
    for path in paths:
        if not path.exists():
            raise FileNotFoundError(path)

    key_sets = [finalize(path) for path in paths]
    if key_sets[0] != key_sets[1]:
        raise ValueError("Reviewer workbooks do not contain the same profiles in the same order")

    for path in paths:
        xml = workbook_xml(path)
        leaked = [marker for marker in FORBIDDEN_MARKERS if marker in xml]
        if leaked:
            raise ValueError(f"{path.name}: private/model fields leaked: {leaked}")
        if "<pane" not in xml:
            raise ValueError(f"{path.name}: freeze panes were not preserved")
        print(f"verified {path}: 200 blinded profiles, validations present, panes frozen, no model-field leakage")


if __name__ == "__main__":
    main()