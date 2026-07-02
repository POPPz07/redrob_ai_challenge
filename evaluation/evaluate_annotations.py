"""Evaluate two completed blind-review workbooks against private model metadata."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.stats import spearmanr
from sklearn.metrics import cohen_kappa_score, ndcg_score


LABEL_COLUMNS = {
    "profile_key": 1,
    "relevance": 14,
    "confidence": 15,
    "risk_flag": 16,
    "rationale": 17,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--reviewer-1", default="outputs/annotation_audit/annotation_reviewer_1.xlsx")
    parser.add_argument("--reviewer-2", default="outputs/annotation_audit/annotation_reviewer_2.xlsx")
    parser.add_argument("--key", default="outputs/annotation_audit/annotation_key.csv")
    parser.add_argument("--output-dir", default="outputs/annotation_audit/evaluation")
    parser.add_argument("--require-complete", action="store_true")
    parser.add_argument(
        "--unlock-holdout",
        action="store_true",
        help="Report holdout metrics only after the development configuration is frozen.",
    )
    return parser.parse_args()


def read_review(path: Path, reviewer: str) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=False, data_only=True)
    sheet = workbook["Annotations"]
    rows = []
    for row_number in range(2, sheet.max_row + 1):
        row = {name: sheet.cell(row=row_number, column=column).value for name, column in LABEL_COLUMNS.items()}
        row["reviewer"] = reviewer
        rows.append(row)
    frame = pd.DataFrame(rows)
    if frame.profile_key.isna().any() or frame.profile_key.duplicated().any():
        raise ValueError(f"{path}: missing or duplicated profile keys")
    return frame


def validate_labels(frame: pd.DataFrame, path: Path) -> list[str]:
    issues: list[str] = []
    for row in frame.itertuples(index=False):
        if pd.isna(row.relevance):
            continue
        if row.relevance not in range(6):
            issues.append(f"{row.profile_key}: relevance must be an integer from 0 to 5")
        if row.confidence not in (1, 2, 3):
            issues.append(f"{row.profile_key}: confidence must be 1, 2, or 3")
        if row.risk_flag not in ("No", "Unsure", "Yes"):
            issues.append(f"{row.profile_key}: risk flag must be No, Unsure, or Yes")
        if not isinstance(row.rationale, str) or len(row.rationale.strip()) < 15:
            issues.append(f"{row.profile_key}: rationale must contain at least 15 characters")
    if issues:
        raise ValueError(f"Invalid labels in {path}:\n" + "\n".join(issues[:20]))
    return issues


def finite_or_none(value: float) -> float | None:
    return float(value) if np.isfinite(value) else None


def dcg_metrics(labels: pd.DataFrame) -> dict[str, float]:
    ordered = labels.sort_values(["weighted_formula_score", "profile_key"], ascending=[False, True])
    relevance = ordered.consensus_relevance.to_numpy(dtype=float)
    model_scores = ordered.weighted_formula_score.to_numpy(dtype=float)
    metrics: dict[str, float] = {}
    for k in (10, 20, 50, 100):
        if len(ordered) >= k:
            metrics[f"ndcg_at_{k}"] = float(ndcg_score([relevance], [model_scores], k=k))
            top = ordered.head(k)
            metrics[f"mean_relevance_at_{k}"] = float(top.consensus_relevance.mean())
            metrics[f"precision_relevance_4plus_at_{k}"] = float((top.consensus_relevance >= 4).mean())
    return metrics


def ranking_metrics(labels: pd.DataFrame) -> dict[str, float | None]:
    ranking: dict[str, float | None] = dcg_metrics(labels)
    correlation = spearmanr(labels.weighted_formula_score, labels.consensus_relevance)
    ranking["spearman_model_vs_consensus"] = finite_or_none(correlation.statistic)
    high_relevance = labels.consensus_relevance >= 4
    selected = labels.current_submission_rank.notna()
    ranking["annotated_4plus_recall_in_current_top100"] = (
        float((high_relevance & selected).sum() / high_relevance.sum()) if high_relevance.any() else None
    )
    ranking["current_top100_mean_annotated_relevance"] = (
        float(labels.loc[selected, "consensus_relevance"].mean()) if selected.any() else None
    )
    return ranking


def main() -> None:
    args = parse_args()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    reviewer_1 = read_review(Path(args.reviewer_1), "reviewer_1")
    reviewer_2 = read_review(Path(args.reviewer_2), "reviewer_2")
    validate_labels(reviewer_1, Path(args.reviewer_1))
    validate_labels(reviewer_2, Path(args.reviewer_2))

    merged = reviewer_1.merge(reviewer_2, on="profile_key", suffixes=("_1", "_2"), validate="one_to_one")
    key = pd.read_csv(args.key)
    if set(merged.profile_key) != set(key.profile_key):
        raise ValueError("Reviewer workbooks and annotation key contain different profile sets")

    complete_1 = int(merged.relevance_1.notna().sum())
    complete_2 = int(merged.relevance_2.notna().sum())
    paired = merged.dropna(subset=["relevance_1", "relevance_2"]).copy()
    completion = {
        "profiles": int(len(merged)),
        "reviewer_1_completed": complete_1,
        "reviewer_2_completed": complete_2,
        "paired_completed": int(len(paired)),
    }
    if args.require_complete and len(paired) != len(merged):
        raise ValueError(f"Reviews are incomplete: {completion}")

    result: dict[str, object] = {
        "warning": "This is an enriched diagnostic sample, not an unbiased estimate over all 100,000 candidates.",
        "completion": completion,
    }

    if len(paired) >= 2:
        paired["relevance_1"] = paired.relevance_1.astype(int)
        paired["relevance_2"] = paired.relevance_2.astype(int)
        paired["consensus_relevance"] = (paired.relevance_1 + paired.relevance_2) / 2
        paired["absolute_disagreement"] = (paired.relevance_1 - paired.relevance_2).abs()
        paired["risk_disagreement"] = paired.risk_flag_1 != paired.risk_flag_2
        paired["needs_adjudication"] = (paired.absolute_disagreement > 1) | paired.risk_disagreement
        paired = paired.merge(key, on="profile_key", validate="one_to_one")

        agreement = {
            "exact_rate": float((paired.absolute_disagreement == 0).mean()),
            "within_one_rate": float((paired.absolute_disagreement <= 1).mean()),
            "mean_absolute_difference": float(paired.absolute_disagreement.mean()),
            "quadratic_weighted_kappa": finite_or_none(
                cohen_kappa_score(paired.relevance_1, paired.relevance_2, weights="quadratic")
            ),
            "risk_exact_rate": float((~paired.risk_disagreement).mean()),
            "adjudication_count": int(paired.needs_adjudication.sum()),
        }
        metric_frame = paired if args.unlock_holdout else paired.loc[paired.evaluation_split == "development"]
        ranking = ranking_metrics(metric_frame)
        ranking_by_split = {
            "development": ranking_metrics(paired.loc[paired.evaluation_split == "development"]),
        }
        if args.unlock_holdout:
            ranking_by_split["holdout"] = ranking_metrics(
                paired.loc[paired.evaluation_split == "holdout"]
            )

        strata = (
            metric_frame.groupby("stratum", as_index=False)
            .agg(
                profiles=("profile_key", "count"),
                mean_human_relevance=("consensus_relevance", "mean"),
                mean_model_score=("weighted_formula_score", "mean"),
                relevance_4plus_rate=("consensus_relevance", lambda values: float((values >= 4).mean())),
                adjudication_rate=("needs_adjudication", "mean"),
            )
            .sort_values("mean_human_relevance", ascending=False)
        )
        strata.to_csv(output_dir / "stratum_metrics.csv", index=False)

        adjudication_columns = [
            "profile_key", "relevance_1", "relevance_2", "confidence_1", "confidence_2",
            "risk_flag_1", "risk_flag_2", "rationale_1", "rationale_2", "absolute_disagreement",
            "stratum", "candidate_id",
        ]
        paired.loc[paired.needs_adjudication, adjudication_columns].sort_values(
            ["absolute_disagreement", "profile_key"], ascending=[False, True]
        ).to_csv(output_dir / "adjudication_queue_private.csv", index=False, quoting=csv.QUOTE_ALL)

        result["holdout_unlocked"] = bool(args.unlock_holdout)
        result["agreement"] = agreement
        result["ranking"] = ranking
        result["ranking_by_split"] = ranking_by_split
        result["strata"] = strata.to_dict(orient="records")

    report_path = output_dir / "human_evaluation_report.json"
    report_path.write_text(json.dumps(result, indent=2, allow_nan=False), encoding="utf-8")
    print(json.dumps(result, indent=2, allow_nan=False))


if __name__ == "__main__":
    main()